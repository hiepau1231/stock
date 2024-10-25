import io
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill
import plotly.graph_objects as go
from plotly.subplots import make_subplots

class PortfolioReportService:
    @staticmethod
    def generate_pdf_report(portfolio, items):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Tiêu đề
        elements.append(Paragraph(f"Báo cáo danh mục {portfolio.name}", styles['Heading1']))
        elements.append(Paragraph(f"Ngày tạo: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Thông tin tổng quan
        total_value = sum(item.current_value for item in items)
        total_cost = sum(item.purchase_value for item in items)
        total_profit = total_value - total_cost

        # Tạo bảng dữ liệu
        data = [['Mã CP', 'Số lượng', 'Giá mua', 'Giá hiện tại', 'Tổng giá trị', 'Lãi/Lỗ', '%']]
        for item in items:
            data.append([
                item.stock.symbol,
                str(item.quantity),
                f"{item.purchase_price:,.2f}",
                f"{item.stock.current_price:,.2f}",
                f"{item.current_value:,.2f}",
                f"{item.profit_loss:,.2f}",
                f"{item.profit_loss_percentage:.2f}%"
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_excel_report(portfolio, items):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Báo cáo danh mục"

        # Tiêu đề
        ws['A1'] = f"Báo cáo danh mục {portfolio.name}"
        ws['A2'] = f"Ngày tạo: {datetime.now().strftime('%d/%m/%Y')}"

        # Header
        headers = ['Mã CP', 'Số lượng', 'Giá mua', 'Giá hiện tại', 'Tổng giá trị', 'Lãi/Lỗ', '%']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="CCCCCC")

        # Data
        for row, item in enumerate(items, 5):
            ws.cell(row=row, column=1).value = item.stock.symbol
            ws.cell(row=row, column=2).value = item.quantity
            ws.cell(row=row, column=3).value = float(item.purchase_price)
            ws.cell(row=row, column=4).value = float(item.stock.current_price)
            ws.cell(row=row, column=5).value = float(item.current_value)
            ws.cell(row=row, column=6).value = float(item.profit_loss)
            ws.cell(row=row, column=7).value = float(item.profit_loss_percentage)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_portfolio_charts(items):
        # Biểu đồ phân bổ danh mục
        fig = make_subplots(
            rows=2, cols=2,
            specs=[[{'type': 'domain'}, {'type': 'bar'}],
                  [{'type': 'scatter', 'colspan': 2}, None]],
            subplot_titles=('Phân bổ danh mục', 'Top holdings', 'Hiệu suất theo thời gian')
        )

        # Pie chart cho phân bổ
        values = [item.current_value for item in items]
        labels = [item.stock.symbol for item in items]
        fig.add_trace(go.Pie(labels=labels, values=values), row=1, col=1)

        # Bar chart cho top holdings
        sorted_items = sorted(items, key=lambda x: x.current_value, reverse=True)[:5]
        fig.add_trace(go.Bar(
            x=[item.stock.symbol for item in sorted_items],
            y=[item.current_value for item in sorted_items],
            name='Giá trị'
        ), row=1, col=2)

        # Line chart cho hiệu suất
        dates = []
        values = []
        for item in items:
            historical_data = item.stock.historical_data.all().order_by('date')
            for data in historical_data:
                date = data.date
                if date not in dates:
                    dates.append(date)
                    values.append(0)
                idx = dates.index(date)
                values[idx] += item.quantity * data.close_price

        fig.add_trace(go.Scatter(
            x=dates,
            y=values,
            mode='lines',
            name='Giá trị danh mục'
        ), row=2, col=1)

        fig.update_layout(
            height=800,
            showlegend=True,
            title_text="Phân tích danh mục đầu tư"
        )

        return fig
